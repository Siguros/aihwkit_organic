# -*- coding: utf-8 -*-

"""MNIST LRTT Transfer Sweep Experiment (Rank & Transfer_every Sweep)

Explores the effect of different rank and transfer_every values.

Configuration:
- Architecture: 784x256x10
- Device: 6T1C A/B tiles (no retention) + Idealized C tile
- Reinit: Orthogonal
- use_onehot: False
- Rank: [4, 16, 64, 128]
- Transfer every: [1, 10, 20, 30, 50, 100, 500, 1000, 5000]
- Transfer LR: 1.0 (fixed)
- Learning rate: 0.1 (fixed)
- Epochs: 30
"""

import os
import sys
import csv
import json
from time import time
from datetime import datetime
import itertools

import torch
from torch import nn
from torch.optim.lr_scheduler import StepLR
from torchvision import datasets, transforms

try:
    import openpyxl
    from openpyxl import Workbook
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False
    print("Warning: openpyxl not installed. xlsx export disabled.")

from aihwkit.nn import AnalogLinear, AnalogSequential
from aihwkit.optim import AnalogSGD
from aihwkit.simulator.configs.lrtt_config import PythonLRTTRPUConfig
from aihwkit.simulator.configs.lrtt_python import PythonLRTTDevice
from aihwkit.simulator.configs import FloatingPointRPUConfig, IOParameters, WeightNoiseType
from aihwkit.simulator.presets.devices import IdealizedPresetDevice, LinearStepDevice
from aihwkit.simulator.rpu_base import cuda

# Check device
USE_CUDA = 0
if cuda.is_compiled():
    USE_CUDA = 1
DEVICE = torch.device("cuda" if USE_CUDA else "cpu")

# Path where the datasets will be stored
PATH_DATASET = os.path.join("data", "DATASET")

# Network architecture
INPUT_SIZE = 784
HIDDEN_SIZE = 256
OUTPUT_SIZE = 10

# Fixed training parameters
EPOCHS = 30
BATCH_SIZE = 128

# Fixed hyperparameters
LEARNING_RATE = 0.1  # Fixed learning rate
TRANSFER_LR = 1.0  # Fixed transfer learning rate

# Hyperparameter search space
RANK_VALUES = [4, 16, 64, 128]  # Rank values to explore
TRANSFER_EVERY_VALUES = [1, 10, 20, 30, 50, 100, 500, 1000, 5000]  # Transfer frequencies to explore

# Output directory for results
RESULTS_DIR = "LRTT_DIRECT_Results"
os.makedirs(RESULTS_DIR, exist_ok=True)


def load_images():
    """Load MNIST dataset."""
    transform = transforms.Compose([
        transforms.ToTensor(),
        transforms.Normalize((0.1307,), (0.3081,))
    ])

    train_set = datasets.MNIST(PATH_DATASET, download=True, train=True, transform=transform)
    val_set = datasets.MNIST(PATH_DATASET, download=True, train=False, transform=transform)
    train_data = torch.utils.data.DataLoader(train_set, batch_size=BATCH_SIZE, shuffle=True)
    validation_data = torch.utils.data.DataLoader(val_set, batch_size=BATCH_SIZE, shuffle=False)

    return train_data, validation_data


def create_model(rank, transfer_every, transfer_lr, burst_limit=None, device_name='cuda'):
    """Create MNIST model with single pulse transfer.

    Args:
        rank (int): Rank for the LRTT decomposition
        transfer_every (int): Transfer frequency
        transfer_lr (float): Transfer learning rate
        burst_limit (int, optional): Transfer burst limit
        device_name (str): Device name ('cuda' or 'cpu')

    Returns:
        nn.Module: Created analog model
    """
    device = torch.device(device_name)

    # 6T1C device (no retention)
    sixt1c_device = LinearStepDevice(
        dw_min=0.001981, up_down=0.0, w_max=1.0, w_min=-1.0,
        gamma_up=-0.1678, gamma_down=0.1410, mult_noise=True,
        dw_min_dtod=0.1, up_down_dtod=0.01, w_max_dtod=0.05, w_min_dtod=0.05,
        gamma_up_dtod=0.05, gamma_down_dtod=0.05, dw_min_std=0.3,
        write_noise_std=0.0182, mean_bound_reference=True,
        lifetime=0.0, lifetime_dtod=0.0, reset=0.0, reset_dtod=0.0,
    )

    # Idealized device (dw_min same as 6T1C, all noise = 0)
    c_device = IdealizedPresetDevice(
        dw_min=0.00198,      # Same as 6T1C
        dw_min_dtod=0.0,      # Device-to-device variation = 0
        up_down_dtod=0.0,     # Already 0 by default
        w_max_dtod=0.0,       # Device-to-device variation = 0
        w_min_dtod=0.0,       # Device-to-device variation = 0
        dw_min_std=0.0        # Cycle-to-cycle variation = 0
    )

    # LRTT device
    device_config_layer1 = PythonLRTTDevice(
        rank=rank, transfer_every=transfer_every, lora_alpha=1.0,
        reinit_gain=0.1, reinit_mode="orthogonal", forward_inject=False,
        use_onehot=False, use_sigma_delta=False,
        unit_cell_devices=[sixt1c_device, sixt1c_device, c_device]
    )

    rpu_cfg_1 = PythonLRTTRPUConfig(device=device_config_layer1)
    io_pars = IOParameters(out_noise=0.006, w_noise_type=WeightNoiseType.NONE)
    rpu_cfg_1.forward = io_pars
    rpu_cfg_1.backward = io_pars

    layer1 = AnalogLinear(784, 256, rpu_config=rpu_cfg_1, bias=True)

    # Set transfer parameters
    layer1.analog_module.controller.transfer_lr = transfer_lr
    if burst_limit is not None:
        layer1.analog_module.controller.transfer_burst_limit = burst_limit

    # Layer 2: FloatingPoint (TRAINABLE, no device noise)
    rpu_cfg_2 = FloatingPointRPUConfig()
    layer2 = AnalogLinear(256, 10, rpu_config=rpu_cfg_2, bias=True)

    model = AnalogSequential(layer1, nn.Sigmoid(), layer2).to(device)

    return model


def train_and_evaluate(model, train_set, val_set, learning_rate):
    """Train the network and return final metrics.

    Args:
        model (nn.Module): Model to be trained
        train_set (DataLoader): Training data
        val_set (DataLoader): Validation data
        learning_rate (float): Learning rate

    Returns:
        dict: Training metrics including final accuracy and loss
    """
    classifier = nn.CrossEntropyLoss()
    optimizer = AnalogSGD(model.parameters(), lr=learning_rate)
    optimizer.regroup_param_groups(model)
    scheduler = StepLR(optimizer, step_size=10, gamma=0.5)

    best_val_acc = 0.0
    final_train_acc = 0.0
    final_val_acc = 0.0
    final_loss = 0.0

    for epoch_number in range(EPOCHS):
        total_loss = 0
        correct_predictions = 0
        total_samples = 0

        model.train()
        for images, labels in train_set:
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            images = images.view(images.shape[0], -1)

            optimizer.zero_grad()
            output = model(images)
            loss = classifier(output, labels)
            loss.backward()
            optimizer.step()

            total_loss += loss.item()
            pred = output.argmax(dim=1, keepdim=True)
            correct_predictions += pred.eq(labels.view_as(pred)).sum().item()
            total_samples += images.size(0)

        # Calculate epoch metrics
        epoch_accuracy = 100. * correct_predictions / total_samples
        avg_loss = total_loss / len(train_set)

        # Validation
        val_accuracy = evaluate(model, val_set)

        scheduler.step()

        # Track best validation accuracy
        if val_accuracy > best_val_acc:
            best_val_acc = val_accuracy

        # Store final epoch metrics
        if epoch_number == EPOCHS - 1:
            final_train_acc = epoch_accuracy
            final_val_acc = val_accuracy
            final_loss = avg_loss

    return {
        'final_train_acc': final_train_acc,
        'final_val_acc': final_val_acc,
        'best_val_acc': best_val_acc,
        'final_loss': final_loss
    }


def evaluate(model, val_set):
    """Evaluate the model on validation set.

    Args:
        model (nn.Module): Model to  | 최고 성능      | Rank=128, Transfer Every=100 → 95.16% evaluate
        val_set (DataLoader): Validation data

    Returns:
        float: Validation accuracy percentage
    """
    correct = 0
    total = 0

    model.eval()
    with torch.no_grad():
        for images, labels in val_set:
            images = images.to(DEVICE)
            labels = labels.to(DEVICE)
            images = images.view(images.shape[0], -1)

            output = model(images)
            _, predicted = torch.max(output.data, 1)

            total += labels.size(0)
            correct += (predicted == labels).sum().item()

    return 100. * correct / total


def run_sweep():
    """Run the complete hyperparameter sweep."""

    print("=" * 80)
    print("MNIST LRTT Transfer Sweep (Rank & Transfer_every Sweep)")
    print("=" * 80)
    print(f"Architecture: {INPUT_SIZE}x{HIDDEN_SIZE}x{OUTPUT_SIZE}")
    print(f"Device: 6T1C A/B + Idealized C")
    print(f"Rank values: {RANK_VALUES}")
    print(f"Transfer every: {TRANSFER_EVERY_VALUES}")
    print(f"Transfer LR (fixed): {TRANSFER_LR}")
    print(f"Learning rate (fixed): {LEARNING_RATE}")
    print(f"Reinit: Orthogonal")
    print(f"use_onehot: False")
    print(f"Epochs: {EPOCHS}")
    print(f"Batch size: {BATCH_SIZE}")
    print(f"Device: {DEVICE}")
    print("=" * 80)
    sys.stdout.flush()

    # Load datasets once
    print("\nLoading MNIST dataset...")
    sys.stdout.flush()
    train_data, validation_data = load_images()
    print(f"Dataset loaded: {len(train_data.dataset)} train, {len(validation_data.dataset)} test")
    sys.stdout.flush()

    # Prepare CSV file for results
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_file = os.path.join(RESULTS_DIR, f"lrtt_direct_sweep_results_{timestamp}.csv")

    # Experiment settings for JSON
    experiment_settings = {
        "experiment_name": "LRTT_DIRECT_Sweep",
        "timestamp": timestamp,
        "architecture": {
            "input_size": INPUT_SIZE,
            "hidden_size": HIDDEN_SIZE,
            "output_size": OUTPUT_SIZE
        },
        "device_config": {
            "ab_tile": "6T1C (LinearStepDevice, no retention)",
            "c_tile": "Idealized"
        },
        "fixed_hyperparameters": {
            "learning_rate": LEARNING_RATE,
            "transfer_lr": TRANSFER_LR,
            "reinit_mode": "orthogonal",
            "reinit_gain": 0.1,
            "lora_alpha": 1.0,
            "use_onehot": False,
            "epochs": EPOCHS,
            "batch_size": BATCH_SIZE,
            "io_out_noise": 0.006,
            "activation": "Sigmoid",
            "layer2_type": "FloatingPoint"
        },
        "sweep_parameters": {
            "rank_values": RANK_VALUES,
            "transfer_every_values": TRANSFER_EVERY_VALUES
        },
        "total_experiments": len(RANK_VALUES) * len(TRANSFER_EVERY_VALUES),
        "device": str(DEVICE)
    }

    # Save experiment settings to JSON
    json_file = os.path.join(RESULTS_DIR, f"experiment_settings_{timestamp}.json")
    with open(json_file, 'w') as f:
        json.dump(experiment_settings, f, indent=4)
    print(f"\nExperiment settings saved to: {json_file}")
    sys.stdout.flush()

    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow([
            'rank', 'transfer_every', 'transfer_lr', 'learning_rate',
            'epochs', 'batch_size', 'use_onehot', 'reinit_mode', 'reinit_gain',
            'lora_alpha', 'io_out_noise', 'activation', 'layer2_type',
            'final_train_acc', 'final_val_acc', 'best_val_acc',
            'final_loss', 'training_time_min'
        ])

    # Run sweep
    total_experiments = len(RANK_VALUES) * len(TRANSFER_EVERY_VALUES)

    print(f"\nTotal experiments: {total_experiments}")
    print("\nStarting sweep...\n")
    sys.stdout.flush()

    experiment_count = 0
    all_results = []
    best_result = None
    best_acc = 0.0

    for rank, transfer_every in itertools.product(RANK_VALUES, TRANSFER_EVERY_VALUES):
        experiment_count += 1

        print(f"\n[{experiment_count}/{total_experiments}] "
              f"rank={rank}, "
              f"transfer_every={transfer_every}, "
              f"transfer_lr={TRANSFER_LR}, "
              f"lr={LEARNING_RATE}")
        sys.stdout.flush()

        # Create model
        device_name = 'cuda' if USE_CUDA else 'cpu'
        model = create_model(
            rank, transfer_every, TRANSFER_LR,
            burst_limit=None, device_name=device_name
        )

        # Train and evaluate
        start_time = time()
        metrics = train_and_evaluate(model, train_data, validation_data, LEARNING_RATE)
        training_time = (time() - start_time) / 60.0

        print(f"  -> Train: {metrics['final_train_acc']:.2f}%, "
              f"Val: {metrics['final_val_acc']:.2f}%, "
              f"Best Val: {metrics['best_val_acc']:.2f}%, "
              f"Time: {training_time:.2f} min")
        sys.stdout.flush()

        # Store result
        result = {
            'rank': rank,
            'transfer_every': transfer_every,
            'transfer_lr': TRANSFER_LR,
            'learning_rate': LEARNING_RATE,
            'epochs': EPOCHS,
            'batch_size': BATCH_SIZE,
            'use_onehot': False,
            'reinit_mode': 'orthogonal',
            'reinit_gain': 0.1,
            'lora_alpha': 1.0,
            'io_out_noise': 0.006,
            'activation': 'Sigmoid',
            'layer2_type': 'FloatingPoint',
            'final_train_acc': metrics['final_train_acc'],
            'final_val_acc': metrics['final_val_acc'],
            'best_val_acc': metrics['best_val_acc'],
            'final_loss': metrics['final_loss'],
            'training_time_min': training_time
        }
        all_results.append(result)

        # Track best result
        if metrics['best_val_acc'] > best_acc:
            best_acc = metrics['best_val_acc']
            best_result = result.copy()

        # Save results with detailed hyperparameters
        with open(csv_file, 'a', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                rank, transfer_every, TRANSFER_LR, LEARNING_RATE,
                EPOCHS, BATCH_SIZE, 'False', 'orthogonal', 0.1,
                1.0, 0.006, 'Sigmoid', 'FloatingPoint',
                metrics['final_train_acc'], metrics['final_val_acc'],
                metrics['best_val_acc'], metrics['final_loss'],
                training_time
            ])

    # Save results to xlsx
    if XLSX_AVAILABLE:
        xlsx_file = os.path.join(RESULTS_DIR, f"lrtt_direct_results_{timestamp}.xlsx")
        wb = Workbook()

        # All results sheet
        ws_all = wb.active
        ws_all.title = "All Results"
        headers = list(all_results[0].keys())
        ws_all.append(headers)
        for result in all_results:
            ws_all.append(list(result.values()))

        # Best result sheet
        ws_best = wb.create_sheet("Best Result")
        ws_best.append(["Parameter", "Value"])
        for key, value in best_result.items():
            ws_best.append([key, value])

        # Summary by rank sheet
        ws_summary = wb.create_sheet("Summary by Rank")
        ws_summary.append(["Rank", "Best Transfer_every", "Best Val Acc (%)", "Final Val Acc (%)"])
        for rank in RANK_VALUES:
            rank_results = [r for r in all_results if r['rank'] == rank]
            best_for_rank = max(rank_results, key=lambda x: x['best_val_acc'])
            ws_summary.append([
                rank,
                best_for_rank['transfer_every'],
                best_for_rank['best_val_acc'],
                best_for_rank['final_val_acc']
            ])

        wb.save(xlsx_file)
        print(f"\nResults saved to xlsx: {xlsx_file}")
    else:
        print("\nWarning: xlsx export skipped (openpyxl not installed)")

    print("\n" + "=" * 80)
    print(f"Sweep completed!")
    print(f"CSV results saved to: {csv_file}")
    print(f"Experiment settings saved to: {json_file}")
    if best_result:
        print(f"\nBest Result:")
        print(f"  Rank: {best_result['rank']}, Transfer_every: {best_result['transfer_every']}")
        print(f"  Best Val Acc: {best_result['best_val_acc']:.2f}%")
    print("=" * 80)
    sys.stdout.flush()


if __name__ == "__main__":
    run_sweep()
